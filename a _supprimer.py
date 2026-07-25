import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, PieChart, Reference

# Create workbook
wb = openpyxl.Workbook()

# Sheet 1: Mode d'emploi & Synthèse
ws_summary = wb.active
ws_summary.title = "Synthèse & Guide"
ws_summary.views.sheetView[0].showGridLines = True

# Sheet 2: Questionnaire BPMN
ws_q = wb.create_sheet(title="Questionnaire BPMN")
ws_q.views.sheetView[0].showGridLines = True

print("Workbook initialized")